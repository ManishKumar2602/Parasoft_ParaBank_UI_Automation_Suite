import openpyxl
import os
from pathlib import Path
 
class ExcelReader:
    @staticmethod
    def read_test_data(file_name, sheet_name="Sheet1"):
        """Read test data from Excel file"""
        base_dir = Path(__file__).parent.parent
        file_path = os.path.join(base_dir, "test_data", file_name)
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Test data file not found: {file_path}")
        
        wb = openpyxl.load_workbook(file_path)
        sheet = wb[sheet_name]
        
        data = []
        headers = []
        
        # Read headers
        for col in range(1, sheet.max_column + 1):
            headers.append(sheet.cell(1, col).value)
        
        # Read data rows
        for row in range(2, sheet.max_row + 1):
            row_data = {}
            for col in range(1, sheet.max_column + 1):
                #make blank value as empty string
                if sheet.cell(row, col).value is None:
                    row_data[headers[col-1]] = ""
                else:
                    row_data[headers[col-1]] = sheet.cell(row, col).value
            data.append(row_data)
        
        return data